import importlib
import matplotlib
import matplotlib.pyplot as plt
import numpy as np
import os
import pandas as pd
import re
import seaborn as sns
import shutil

from datetime import timedelta
from file_read_backwards import FileReadBackwards

from functools import partial
from getpass import getuser
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from pathlib import Path
from prettytable import PrettyTable
from scipy.fft import fft, fftfreq
from socket import gethostname
from statistics import stdev, mean
from tqdm import tqdm
from warnings import warn

import utils.constants as cst
import utils.fetch as fch
import utils.find as fd
import utils.misc as misc
import utils.plot as pl
import utils.classes as cl

# Args for the figures
matplotlib.rcParams['xtick.major.width'] = 1
matplotlib.rcParams['ytick.major.width'] = 1
matplotlib.rcParams['lines.markeredgewidth'] = 1
matplotlib.rcParams['xtick.labelsize'] = 12
matplotlib.rcParams['ytick.labelsize'] = 12
matplotlib.rcParams['axes.linewidth'] = 1

# Backend
matplotlib.use('QtAgg')

# Use seaborn to generate plots
sns.set()
sns.set_style("whitegrid")

# %% ===================================================================================================

def reload():
    module = importlib.import_module(__name__)
    importlib.reload(module)
    importlib.reload(cst)
    importlib.reload(fd)
    importlib.reload(fch)
    importlib.reload(misc)
    importlib.reload(pl)
    importlib.reload(cl)
    
    print('Reloaded.')

# * ===================================================================================================

def plot_data(runs, *, specdir: str, probe: str = None, freq: bool = False, **kwargs):
    # Gather runs, run data, and CSV DataFrame
    run_dirs, run_nb, csv_df = pl.gather_runs(runs)
    misc.print_header(run_dirs)
    run_pp_df_list = pl.gather_data(run_dirs, specdir, probe, **kwargs)
    
    # Set up the figure parameters
    ax = pl.set_figure_params(probe, specdir)
    handle_prefix = "Probe " if probe is not None else ""
    unit = None
    
    # Iterate over each processed run DataFrame
    for data in run_pp_df_list:
        run_id, pp_dir, df = data.values()
        frmt_legend = " | " + run_id if run_nb > 1 else ""

        if freq:
            sampling_rate = len(df["Time"]) / df["Time"].iloc[-1]
            pl.set_axis_labels(ax, freq=True)
            print(f"\n{cst.bmag}------------\n# FFT {run_id}\n------------{cst.reset}\n")
            
            # Plot frequency data and yield handles
            handles = [h for h in pl.plot_freq_data(run_id, ax, df, handle_prefix, frmt_legend, sampling_rate, **kwargs)]
        else:
            unit = fch.fetch_unit(df=df, pp_dir=pp_dir, probe=probe, csv_df=csv_df, **kwargs)
            pl.set_axis_labels(ax, unit=unit)
            
            # Plot time data and yield handles
            handles = [h for h in pl.plot_time_data(ax, df, handle_prefix, frmt_legend)]

    if unit is None:
        plt.gca().set_ylabel(None)
    
    # Format the legend and set the title if specified
    pl.format_legend(ax, handles)
    if 'title' in kwargs:
        title = kwargs.get('title')
        ax.set_title(title, fontsize=20)

    print('\nDisplaying the figure...\n')
    
    # Adjust figure layout and display the figure
    plt.tight_layout()
    plt.show()

# * ========================= PARTIAL FUNCTIONS =========================

plot_probes = partial(plot_data, specdir=None)
plot_residuals = partial(plot_data, specdir=None, probe=None)

# * ===================================================================================================

# TODO BAR CHART ===================================================================================================

# def bar_chart(target, *,
#               rng: int = cst.RNG,
#               specdir: str = None,
#               probe: str = None,
#               **kwargs) -> None:
    
#     # ! Mutual exclusion of arguments
#     if probe != None and specdir != None:
#         raise ValueError('probe and specdir are mutually exclusive.')

    
#     # Find the run(s) path
#     run_dirs = []
#     if isinstance(target, str):
#         target = [target]
#     for tar in target:
#         run_dirs += _find_runs(tar)

#     # ! If no run found
#     if len(run_dirs) == 0:
#         raise ValueError(f"No run directory found with this list.")
    
#     # Else, print the table showing the project and runs
#     else:
#         _print_header(run_dirs)
        
#         # If no probe, legend not modified
#         if probe == None:
#             lgd = ''
#             sig_list = _get_sig_list(run_dirs,
#                                         specdir=specdir,
#                                         **kwargs)
            
#         # If probe, "Probe #" added to the legend 
#         else:
#             lgd = f'Probe{space}'
#             sig_list = _get_sig_list(run_dirs,
#                                         probe=probe,
#                                         **kwargs)  

#         # Initialization
#         handles, xlabel = [], []
#         df_mean = pd.DataFrame()
#         _, ax = plt.subplots(figsize=(12, 27/4))
#         run_number = len({sig['run_id'] for sig in sig_list})
#         pp_dir_number = len({sig['pp_dir'] for sig in sig_list})
#         xpos = np.arange(len(sig_list))
#         ax.ticklabel_format(axis='y', style='sci', scilimits=(-1, 0))
        
#         # Loop over the datasets to be plotted
#         for sig in sig_list:
#             run_id = sig['run_id'] 
#             pp_dir = sig['pp_dir']
#             df = sig['df'].iloc[:, 1:] # remove time column
            
#             # Initialize a dict representing the mean data for a run/pp_dir combination
#             mean_dict = {'run': run_id, 'pp_dir': pp_dir}
#             mean_dict.update({col: df[col].tail(rng).mean() for col in df.columns})
            
#             # Add a new row of mean values in df_mean for each run/pp_dir combination
#             df_mean = pd.concat([df_mean, pd.Series(mean_dict).to_frame().T])
                            
#             # Format strings for LateX font
#             format_string = lambda x: x.replace('_', underscore).replace(' ', space)
#             frmt_run = format_string(run_id)
#             frmt_pp_dir = format_string(pp_dir)
            
#             # Format xlabel with pp_dir and/or run_id
#             if run_number > 1 and pp_dir_number > 1:  
#                 frmt_legend = f'{marker}{frmt_pp_dir}{sep}run{frmt_run}{marker}'
#             elif pp_dir_number > 1:
#                 frmt_legend = f'{marker}{frmt_pp_dir}{marker}'
#             else:
#                 frmt_legend = f'{marker}run{frmt_run}{marker}'
#             xlabel.append(frmt_legend)
            
#         # Set a multi index with the run/pp_dir combination
#         df_mean = df_mean.set_index(['run', 'pp_dir'])

#         # Set the width of the rectangles
#         width = 1.5 / (len(xpos) * len(df.columns))

#         # Loop over the columns to plot each series of data with their handle
#         for i, col in enumerate(df_mean.columns):
#             handle = f'{marker}{lgd + col}{marker}'
#             handles.append(handle)
#             rect = ax.bar(xpos + i * width,
#                             df_mean[col],
#                             width = width,
#                             label = handle)
#             if 'fancyplot' in kwargs or 'fp' in kwargs:
#                 frmt = '${:.2e}$'
#             else:
#                 frmt = '{:.2e}'
#             ax.bar_label(rect, padding=3, fmt=frmt)
            
#         # Plot parameters
#         ax.legend(loc='upper center',
#                     bbox_to_anchor = [0.5, -0.1],
#                     framealpha = 1,
#                     frameon = False,
#                     ncol = _ncol(handles),
#                     borderaxespad=0,
#                     fontsize=12)
        
#         # If a unit is specified for the y axis
#         if 'unit' in kwargs:
#             ax.set_ylabel(f'{marker}{kwargs.get("unit")}{marker}', labelpad=10)

#         # If a title is specified
#         if 'title' in kwargs:
#             title = format_string(kwargs.get('title'))
#             ax.set_title(f'{marker}{title}{marker}', fontsize=20, fontweight='cst.bold')
            
#         # Set the xticks at the center of the grouped rectangles
#         ax.set_xticks(xpos + width * (len(df_mean.columns) - 1) / 2, xlabel, fontsize=15)
        
#         # Verbose
#         print('\nDisplaying the figure...\n')
        
#         plt.tight_layout()
#         plt.show()

# TODO BAR CHART ===================================================================================================

# * ===================================================================================================

# TODO PLOT TIME ===================================================================================================

# def plot_time(target, *, x='iterations', skipstart=10, **kwargs):
#     col1 = cst.HOPIUM_PALETTE['hopium']
#     col2 = cst.HOPIUM_PALETTE['red']

#     if x == 'iterations' or x == 'time':
#         fig, ax = plt.subplots(figsize=(12, 27/4))
#         ax_cumul = ax.twinx()
#     elif x =='both':
#         fig, (ax_time, ax_iter) = plt.subplots(1, 2, figsize=(12, 27/4), sharey=False)
#         ax_time_cumul = ax_time.twinx()
#         ax_iter_cumul = ax_iter.twinx()
    
#     # Label formatting
#     marker, space, underscore, sep = _display_settings(**kwargs)
#     # Find the run(s) path
#     run_dirs = []
#     log_files = []
#     if isinstance(target, str):
#         target = [target]
#     for tar in target:
#         run_dirs += _find_runs(tar)

#     # ! If no run found
#     if len(run_dirs) == 0:
#         raise ValueError(f"No run directory found with this list.")
    
#     # If at least one run found
#     else:
#         time_pattern = compile(r'Time = ([\d.]+)s')
#         exec_pattern = compile(r'ExecutionTime\s*=\s*([\d.]+)\s*s')
#         current_iter = 0
#         prev_iter = 0
#         restarts = []
#         cumul_time = []
#         _print_header(run_dirs)
#         for run in run_dirs:
#             run_id = search('(?<=run)\d{3}\w*', run).group(0)
#             print(f'\nRun {_cst.bmag}{run_id}{_cst.reset}')
#             log_files += _find_logs(run)
#             data_iter = {'timestep': [], 'exec_time': []}

#             for log in log_files:
#                 print('Log: ' + log)
#                 if len(data_iter['exec_time']) > 0:
#                     current_iter = data_iter['exec_time'][-1]

#                 with open(log, 'r') as f:
#                     time_bool = False
#                     for line in f:
#                         # Patterns to find the lines starting with 'Time' and 'ExecutionTime' and extract the values
#                         time_match = match(time_pattern, line)
#                         exec_match = match(exec_pattern, line)
#                         # If a line starts with 'Time' and time_bool is False
#                         if time_match and not time_bool:
#                             # Extract the float value of Time
#                             time_value = float(time_match.group(1))
#                             # Set time_bool to True because a line starting with 'Time' has been found
#                             time_bool = True
#                         # If a line starts with 'ExecutionTime' and time_bool is True
#                         elif exec_match and time_bool:
#                             # The value of the previous execution time is kept to calculate the time difference
#                             prev_iter = current_iter
#                             # The current execution is extracted
#                             current_iter = float(line.split()[2])
#                             # Set time_bool to False to cst.reset the boolean indicating if a 'Time' line has been found
#                             time_bool = False
#                             # The 'Time' value is added to the timestep list
#                             data_iter['timestep'].append(time_value)
#                             # If the iteration is not the first one
#                             if len(data_iter['exec_time']) > 0:
#                                 # The timestep duration is calculated and added to the exec_time list
#                                 data_iter['exec_time'].append(current_iter - prev_iter)
#                                 cumul_time.append(cumul_time[-1] + current_iter - prev_iter)
#                             # If it is the first iteration
#                             else:
#                                 # The duration of the first timestep is added to the exec_time list
#                                 data_iter['exec_time'].append(current_iter)
#                                 cumul_time.append(current_iter)
#                 restarts.append((data_iter['timestep'][-1], len(data_iter['timestep'])))
#             ax_time.scatter(data_iter['timestep'][skipstart:],
#                             data_iter['exec_time'][skipstart:],
#                             color=col1, marker='.', s=1.5)
            
#             ax_time_cumul.plot(data_iter['timestep'][skipstart:],
#                             cumul_time[skipstart:], color=col2)
            
#             ax_iter.scatter(np.arange(skipstart, len(data_iter['timestep'])),
#                             data_iter['exec_time'][skipstart:],
#                             color=col1, marker='.', s=1.5)
            
#             ax_iter_cumul.plot(np.arange(skipstart, len(data_iter['timestep'])),
#                             cumul_time[skipstart:], color=col2)

#     for r in [element[0] for element in restarts[:1]]:
#         ax_time.axvline(r, linestyle=':')
#     for p in [element[1] for element in restarts[:1]]:
#         ax_iter.axvline(p, linestyle=':')

#     ax_time.set_xlabel(f'{marker}Time{space}(s){marker}', labelpad=10)
#     ax_iter.set_xlabel(f'{marker}Iterations{marker}', labelpad=10)

#     ax_time.tick_params(labelleft=True, labelright=False,  left=True, right=False)
#     ax_time_cumul.tick_params(labelleft=False, labelright=False,  left=False, right=True)
#     ax_iter.tick_params(labelleft=False, labelright=False,  left=True, right=False)
#     ax_iter_cumul.tick_params(labelleft=False, labelright=True,  left=False, right=True)

#     for ticklabel in ax_time.get_yticklabels():
#         ticklabel.set_color(col1)
#     for ticklabel in ax_iter_cumul.get_yticklabels():
#         ticklabel.set_color(col2)

#     fig.text(-0.03, 0.35, f'{marker}Time{space}per{space}Iteration{space}(s){marker}',
#              fontsize=15, rotation=90, color='#ef476f')
#     fig.text(1, 0.42, f'{marker}Total{space}Time{space}(s){marker}',
#              fontsize=15, rotation=-90, color='#0096c7')
#     fig.tight_layout()
#     plt.show()

# TODO PLOT TIME ===================================================================================================

# * ===================================================================================================

def sim_time(run):

    # Get run and log files  
    run_list = fd.find_runs(run)

    # ! Wrong run arg
    if not run_list:
        raise ValueError('No run path found.')
    
    log_files = fd.find_logs(run_list[0])
    times_list = []

    # Parsing log files to find the line containing the last timestep
    for log in log_files: 
        with FileReadBackwards(log, encoding='utf-8') as frb:
            for line in frb:
                # Steady simulations
                if line.startswith('ExecutionTime') and misc.issteady(run):
                    stime = int(line.split()[-2])
                    break
                # Unsteady simulations
                elif line.startswith('Time') and not misc.issteady(run):
                    stime = np.float64(line.split()[-1][:-1])
                    break
        times_list.append(stime)
    
    # Total time calculation
    total_time = sum(times_list)
    return total_time

# * ===================================================================================================

def stop_sim(run):
    """
    This function stops the simulation by modifying the 'stopAt' line in the controlDict file
    of the specified run. It finds the path to the controlDict file and opens it, 
    then it replaces the first occurrence of 'endTime' with 'writeNow' and writes the modified line
    to the temporary file. Finally, the script uses the os.replace function to replace the original
    file with the temporary file.
    
    Parameters:
    - run (str): The name of the run folder containing the controlDict file
    
    Returns:
    - None
    """
    run_path = fd.find_runs(run)[0]
    # Path to the controlDict file in the run
    controlDict_path = fd.find_files("controlDict", root_dir=run_path + '/system/')[0]
    # Copy of the controlDict file where the new line will be written
    temp_file_path = controlDict_path + ".tmp"
    # Read the controlDict file to find the line starting with "stopAt"
    with open(controlDict_path, "r") as f, open(temp_file_path, "w") as temp:
            for line in f:
                if line.startswith("stopAt"):
                    # Replace "endTime" by "writeNow" in the temp file 
                    temp.write(line.replace("endTime", "writeNow"))
                else:
                    temp.write(line)
    print('Simulation stopping...')

    # Replace the original controlDict file by the temp file
    os.replace(temp_file_path, controlDict_path)

# * ===================================================================================================

def recap_sim(runs: str, *,
              geometry_name: str = None) -> None:

    run_paths = fd.find_runs(runs)
    new_rows = pd.DataFrame()

    for run_path in run_paths:
        run_id = run_path.name
        steady = misc.issteady(run_id)
        
        # Find the path to the run directory (assuming the run directory is unique)
        date_check = False
        n_procs_check = False
        date = ''
        n_procs = ''
        turbulence_model = ''
        
        # Get the turbulence model from the "turbulenceProperties" file
        momentumTransport_path = run_path / "constant" / "momentumTransport"
        with open(momentumTransport_path, 'r') as f:
            for line in f:
                if 'model' in line:
                    turbulence_model = line.split()[-1].strip(";")
                    break

        # Get the number of iteratsions from the "controlDict" file
        log_files = sorted(fd.find_logs(run_path))
        
        data_dict = {
            'Project': [run_path.parent.name],
            'Run': [run_id],
            'User': [getuser()],
            'Workstation': [gethostname()],
            'Geometry': [geometry_name],
            'Clock Time': [str(timedelta(seconds=int(sim_time(run_id))))],
            'Turbulence Model': [turbulence_model],
        }

        for i, log in enumerate(log_files):
            if i == 0:
                with open(log, 'r') as f:
                    for line in f:
                        if line.startswith('Date'):
                            date = ' '.join(line.split()[-3:])
                            date_check = True
                        elif line.startswith('nProcs'):
                            n_procs = int(line.split()[-1])
                            n_procs_check = True
                        if date_check and n_procs_check:
                            break
            if i == len(log_files) - 1:
                with FileReadBackwards(log) as frb:
                    if steady:
                        for line in frb:
                            if line.startswith("Time ="):
                                data_dict['Iterations'] = int(line.split()[-1])
                                break
                    else:
                        for line in frb:
                            if line.startswith("Time ="):
                                data_dict['Simulated Time (s)'] = np.float64(line.split()[-1][:-1])
                                break
        
        data_dict.update({'Date': [date], '# Procs': [n_procs]})

        # Add the data to the dataframe
        df = pd.DataFrame.from_dict(data_dict)

        # Move column date to the first position
        cols = df.columns.tolist()
        cols.insert(0, cols.pop(8))
        df = df[cols].round(3)
        new_rows = pd.concat([new_rows, df], ignore_index=True)

    pt = PrettyTable()

    for col in new_rows.columns:
        pt.add_column(col, new_rows[col].values)
        pt.align[col] = 'c'

    print(pt)
    
    
        # # Get data from run
        # run_pp_df_list += [data for data in _get_data_from_run(run_path,
        #                                                           specdir = specdir,
        #                                                           probe = probe,
        #                                                           **kwargs)]
        # if not run_pp_df_list:
        #     raise ValueError("No data found with such input.")
        
        # mean_dict = {}
        # for sig in sig_list:
        #     df = sig['df']
        #     # Remove the Time column
        #     df = df.iloc[:,1:]
        #     # Initialize a dict representing the mean data for a run/pp_dir combination
        #     mean_dict.update({col: df[col].tail(rng).mean() for col in df.columns})
        #     mean_dict = {(probe + col) if col[0].isdigit() else col: value for col, value in mean_dict.items()}

        #     data_dict.update(mean_dict)

        
        
        # new_rows = pd.concat([new_rows, df], axis=0)
        
        
        # new_cols = new_rows.columns

# TODO =====================================================================================================

# def update_excel(df: pd.DataFrame, xl_path: str = '/home/victorien/ofpostpro/recap_sim.xlsx') -> None:
        
#     if os.path.isfile(xl_path):
#         existing_rows = pd.read_excel(xl_path)
#         total_rows = existing_rows
#         existing_cols = existing_rows.columns
        
#     # ! Wrong path
#     else:
#         raise ValueError('The path to the Excel file does not exist.')

#     # Find columns in new_rows that are missing in existing_df
#     missing_cols_xl = list(set(new_cols) - set(existing_cols))
#     for col in missing_cols_xl:
#         total_rows[col] = ''
#     if missing_cols_xl:
#         print('\nNew column(s) added:',
#               f'{cst.bmag}{f"{cst.reset}, {cst.bmag}".join(sorted([str(i) for i in missing_cols_xl]))}{cst.reset}.')
    
#     # Find columns in new_rows that are missing in new_rows
#     missing_cols_new_rows = list(set(existing_cols) - set(new_cols))
#     for col in missing_cols_new_rows:
#        new_rows[col] = ''

#     # Append the new rows to the existing DataFrame
#     appended_df = pd.concat([total_rows, new_rows], ignore_index=True) 
#     sorted_missing_cols = sorted(missing_cols_xl)
#     sorted_cols = existing_cols.to_list() + sorted_missing_cols
#     appended_df = appended_df[sorted_cols]
#     # Save the updated DataFrame to a new Excel file
#     appended_df.to_excel(xl_path, index=False)
    
# %%

#     _format_excel(xl_path)