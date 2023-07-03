# %% ---------------------------------------------bold-----------------------------------------------------------## %% --------------------------------------------------------------------------------------------------------#import os

import os
import re
import math
import shutil
import matplotlib
import warnings
import pandas as pd
import constants as cst

from tqdm import tqdm
from functools import partial
from prettytable import PrettyTable
from file_read_backwards import FileReadBackwards
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# ANSI escape sequences of different colors for text output
reset = "\033[0m"
bold = "\033[1m"
# Colors
red = "\033[0;91m"
green = "\033[0;92m"
yellow = "\033[0;93m"
blue = "\033[0;94m"
magenta = "\033[0;38;5;199m"
cyan = "\033[0;96m"
white = "\033[0;97m"
# Bold colors
bgray = "\033[1;90m"
bred = "\033[1;91m"
bgreen = "\033[1;92m"
bblue = "\033[1;94m"
bmag = "\033[1;38;5;199m"
bcyan = "\033[1;96m"

def _find_logs(run_path: str) -> list:
    """
    Given a path to a directory, returns a list of all log files in the directory and its 'logs' subdirectory.

    Args:
        run_path (str): The path of the directory to search for log files.

    Returns:
        list: A list of all log files in the directory and its 'logs' subdirectory.
    """
    # Initialize an empty list to store the log files.
    log_files: list = []
    filtered_log_files: list = []
    pattern: re.Pattern = re.compile(cst.LOG_REGEX)

    # Search for log files in the run_path directory.
    log_files += [f.path for f in os.scandir(run_path) 
                  if re.search(pattern, f.name)]

    # Check if the 'logs' subdirectory exists in the run_path directory.
    if os.path.isdir(os.path.join(run_path, 'logs')):
        # If the 'logs' subdirectory exists, search for log files in it.
        log_files += [f.path for f in os.scandir(os.path.join(run_path, 'logs')) 
                      if re.search(pattern, f.name)]

    # Remove the 'log.potentialFoam' file from the log_files list if it exists.
    filtered_log_files = [log for log in log_files if os.path.basename(log) != 'log.potentialFoam']

    # Return the log_files list.
    return filtered_log_files

# %% ===================================================================================================

def _issteady(run: str) -> bool:
    log_file = _find_logs(_find_runs(run)[0])[0]
    with FileReadBackwards(log_file) as frb:
        for line in frb:
            if line.startswith("Time ="):
                if line.split()[-1].isdigit():
                    return True
                else:
                    return False
                
# %% ===================================================================================================

def _csv_postpro_to_df(csv_file: str = '/home/victorien/ofpostpro/postpro_directories.csv') -> pd.DataFrame:
    """
    Convert a CSV file containing post-processing directories and labels into a Pandas DataFrame.

    Args:
        csv_file (str): Path to the CSV file.

    Returns:
        pandas.DataFrame: A DataFrame with hierarchical indexing by directory and label type.
    """

    # Convert the csv file into a DataFrame while filling the empty cells with ""
    csv_df: pd.DataFrame = pd.read_csv(csv_file, sep=';').fillna(pd.NA)

    # Fill missing directory values using the most recent non-null value
    csv_df['Directories'] = csv_df['Directories'].fillna(method='ffill')

    # Set hierarchical indexing by directory and label type
    csv_df = csv_df.set_index(['Directories', 'Label Type'])

    return csv_df

# %% ===================================================================================================

def _ncol(handles: list) -> int:

    max_text_length = 60
    nhandles = len(handles)
    total_length = sum(len(text) for text in handles) + 3 * nhandles
    
    if total_length > max_text_length:
        if nhandles > 6:
            ncol = 4
        else:
            ncol = max(int(nhandles / 2), int((nhandles + 1) / 2))
        
        row_index = range(0, nhandles - ncol, ncol)
        for i in row_index:
            words_in_row = [handles[k] for k in range(i, i + ncol)]
            if sum(len(word) for word in words_in_row) > max_text_length:
                ncol -= 1
                break
    else:
        ncol = nhandles
    return ncol
    
# %% ===================================================================================================

def _get_postpro_labels(database: pd.DataFrame, 
                        directory: str, 
                        category: str) -> list:
    """
    Get a list of post-processing labels for a given directory and category.

    Args:
        df (pandas.DataFrame): DataFrame containing post-processing directories and labels.
        directory (str): Directory to search for labels.
        category (str): Category of labels to retrieve, 'in_file' or 'postpro'.

    Returns:
        List[str]: List of labels for the given directory and category.
    """
    # Use hierarchical indexing to retrieve all labels for the given directory and category
    labels: pd.Series = database.loc[(directory, category), :]

    # Filter out any NaN values from the resulting Series
    filtered_labels: list = [c for c in labels.to_numpy() if not pd.isna(c)]

    return filtered_labels

# %% ===================================================================================================

def _label_names(fpath: str) -> dict:
    if 'probes' in fpath or 'residuals' in fpath:
        # If the file is a probes or residuals file, the labels are determined based
        # on the format of the file.
        with open(fpath, 'r') as f:
            if fpath.endswith('U'):
                # If the file is a velocity file, the labels are composed of the time
                # and the velocity components.
                for line in f:
                    if line.startswith('# Time'):
                        file_labels = line.strip().split()[2:]
                        postpro_labels = ['Time'] + [pnum + i for pnum in file_labels for i in (' - Ux', ' - Uy', ' - Uz')]
                        break
            else:
                # If the file is not a velocity file, the labels are composed of the
                # time and the variables in the file.
                for line in f:
                    if line.startswith('# Time'):
                        file_labels = line.strip().split()[1:]
                        postpro_labels = file_labels
                        break
    else:
        # If the file is not a probes or residuals file, the labels are obtained from
        # a CSV file that maps the directories and label types to the corresponding labels.
        postpro_dir: list = fpath.split('/')[-3]
        csv_df: pd.DataFrame = _csv_postpro_to_df()
        
        file_labels: list = _get_postpro_labels(
            database = csv_df,
            directory = postpro_dir,
            category = 'in_file',
        )
        postpro_labels: list = _get_postpro_labels(
            database = csv_df,
            directory = postpro_dir,
            category = 'postpro',
        )

    # Return a dictionary containing the original file labels and the post-processing labels.
    return {
        'file_labels': file_labels,
        'postpro_labels': postpro_labels,
    }

# %% ===================================================================================================
    
def _find_paths(target: str, *,
                dtype: str, 
                root_dir: str = cst.DEFAULT_DIR,
                **kwargs) -> list:
    
    # ============================= VARIABLES =============================
    
    target_pattern = re.compile(target)
    stack = [os.path.abspath(root_dir)]
    output_paths: list = []
    pattern_excluded_dir = re.compile(cst.EXCLUDED_REGEX)
    pattern_postpro_dir = re.compile(cst.POSTPRO_REGEX)
 
    # =============================== START ===============================

    while stack:
        dirpath: str = stack.pop()

        # Find a dir
        if dtype == 'dir':

            # Filter items to search
            filtered_items = [
                e for e in os.scandir(dirpath) # Item in dirpath
                if e.is_dir() # Item is a dir
                and e.name not in cst.EXCLUDED_ITEMS # Item's name not in excluded list
                and not bool(re.search(pattern_excluded_dir, e.name)) # Item's name not in excluded pattern 
            ]

            # Determine if the entry goes to the output paths or to the stack
            for entry in filtered_items:
                is_match = re.search(target_pattern, entry.name)
                if bool(is_match):
                    output_paths.append(entry.path)
                else:
                    stack.append(entry.path)
                    
        # Find a file
        elif dtype == 'file':

            # Filter items to search
            filtered_items: list = [
                e for e in os.scandir(dirpath) # Item in dirpath
                if e.name not in cst.EXCLUDED_ITEMS # Item's name not in excluded list
                and bool(re.search(pattern_postpro_dir, e.path)) # Item's name not in excluded pattern 
            ]

            # Determine if the entry goes to the output paths or the stack
            for entry in filtered_items:

                # If the entry is a file
                if entry.is_file():

                    # If the entry name matches the target append to output paths
                    if bool(re.search(target_pattern, entry.name)):
                        output_paths.append(entry.path)

                    # Else, get the dict of labels of the file
                    elif entry.name.endswith('.dat'):
                        label_dict = _label_names(entry.path)

                        # If a specific label is searched in the file
                        if 'search' in kwargs:
                            to_search = kwargs.get('search')
                            label_list = label_dict.get('postpro_labels')
                            lab_is_match = bool({lab for lab in label_list if bool(re.search(to_search, lab))})
                            if lab_is_match:
                                output_paths.append(entry.path)
                        
                        else:
                            output_paths.append(entry.path)

                # If item is a dir, add it to the stack
                else:
                    stack.append(entry.path)
    
    return sorted(output_paths)

# * ========================= PARTIAL FUNCTIONS =========================

_find_dirs = partial(_find_paths, dtype='dir')
_find_files = partial(_find_paths, dtype='file')

# * ===================================================================================================

def _find_runs(target: str, *,
               root_dir: str = cst.DEFAULT_DIR,
               **kwargs) -> list:
    
    # Find all the dirs in the root dir
    all_dirs =  _find_dirs(target, root_dir=root_dir, **kwargs)
    # Keep the dirs containing a 'system' dir and a 'constant' dir
    run_dirs = [d for d in all_dirs if os.path.isdir(os.path.join(d, 'system')) and os.path.isdir(os.path.join(d, 'constant'))]

    return run_dirs

# * ===================================================================================================

def _concat_data_files(file_paths: list) -> dict:
        
        data_list = []

        # Sort file paths 
        if all(f.split('/')[-2].isdigit() for f in file_paths):
            timesteps = [int(f.split('/')[-2]) for f in file_paths]
        else:
            timesteps = [float(f.split('/')[-2]) for f in file_paths]
        sorted_file_paths = [x for _, x in sorted(zip(timesteps, file_paths), key=lambda pair: pair[0])]

        # Get columns name
        cols = _label_names(sorted_file_paths[0])['postpro_labels']

        # Loop over the file paths
        for fpath in sorted_file_paths:
            pp_dir = fpath.split('/')[-3]
            timestep = fpath.split('/')[-2]

            # Parse file
            with open(fpath, 'r') as f:
                for line in f:
                    if line.startswith('# Time'):
                        break

                # Verbose
                if 'probe' in fpath:
                    print(f'Parsing {bmag}probe {os.path.basename(fpath)}{reset} at timestep {bmag}{timestep}{reset}:')
                else:
                    print(f'Parsing {bmag}{pp_dir}{reset} at timestep {bmag}{timestep}{reset}:')
                
                # Set progress bar
                pbar = tqdm(total=0, unit=' lines')
                for line in f:
                    data = tuple(line.split())
                    data_list.append(data)
                    pbar.update(1)
                pbar.close()

                #Verbose
                print('File parsed.')
        
        # If at least 2 files are concatenated, display their start timestep
        if len(file_paths) > 1:
            fmt_sep = f"{reset}, {bmag}"
            fmt_timesteps = f"{fmt_sep}".join(sorted([str(i) for i in timesteps]))
            print(f'Concatenated files at timesteps {bmag}{fmt_timesteps}{reset}.')

        return data_list, cols

# * ===================================================================================================

def _remove_NaN_columns(concat_data: tuple) -> pd.DataFrame:

    data, headers = concat_data
        
    # Check for NaN columns to remove
    # Initialize the indexes of headers and columns to drop
    idx_headers_to_drop = []
    idx_cols_to_drop = []

    # DataFrame created without the headers
    df = pd.DataFrame(data)
    
    # Loop over the columns of the DataFrame
    for i in tqdm(range(len(df.columns)),
                    bar_format=cst.BAR_FORMAT,
                    ascii=' |',
                    colour='green',
                    desc='Checking for NaN columns'):
        
        # If a column filled with NaN is found    
        if df.iloc[:, i][df.iloc[:, i] != 'N/A'].size == 0:

            # Its index is saved to remove the corresponding header and column
            idx_cols_to_drop.append(i)
            idx_headers_to_drop.append(i)

            # If the velocity vector is NaN, 
            # the Ux, Uy and Uz headers must be removed
            if 'Uy' in headers[i+1] or 'Uz' in headers[i+1]:
                idx_headers_to_drop.append(i+1)
                if 'Uz' in headers[i+2]:
                    idx_headers_to_drop.append(i+2)

    # Verbose
    if idx_cols_to_drop:
        print('NaN column(s) removed.')
    else:
        print('No column to remove.')

    # Associate the columns with their headers
    df = df.drop(idx_cols_to_drop, axis=1)
    headers = [val for i, val in enumerate(headers) if i not in idx_headers_to_drop]
    df.columns = headers
    
    return df 

# * ===================================================================================================
    
def _remove_NaN_cells(df: pd.DataFrame) -> pd.DataFrame:

    # Check for NaN cells to remove
    tqdm.pandas(bar_format=cst.BAR_FORMAT,
                ascii=' |',
                colour='green',
                desc='Checking for remaining NaN cells')
    
    # Save the initial number of rows in the DataFrame
    previous_size = df.size

    # Drop the rows with one or more NaN or None cells
    df = df[~df.progress_apply(lambda x: x.isin(['N/A', None])).any(axis=1)]

    # Verbose
    if df.size != previous_size:
        print('NaN cell(s) removed.')
    else:
        print('No cell to remove.')

    return df

# * ===================================================================================================

def _remove_parenthesis(df: pd.DataFrame) -> pd.DataFrame:

    # Check for parenthesis to remove
    cols_with_parenthesis = [col for col in df.columns if df[col][1].startswith('(') or df[col][1].endswith(')')]

    # If one or more columns contain cells with parenthesis
    if len(cols_with_parenthesis) > 0:

        # Progress bar setup
        tqdm.pandas(bar_format=cst.BAR_FORMAT,
                    colour='green',
                    ascii=' |',
                    desc='Removing parenthesis')
        
        # Remove the parenthesis in the identified columns
        df[cols_with_parenthesis] = df[cols_with_parenthesis].progress_apply(lambda x: x.str.replace(r'[()]', '', regex=True))

        # Verbose
        print('Parenthesis removed.')

    return df

# * ===================================================================================================

def _filter_data(df: pd.DataFrame, **kwargs) -> pd.DataFrame:

    # If one or more columns are specified by the user
    if 'usecols' in kwargs:
        usecols = kwargs.get('usecols')
        if isinstance(usecols, str):
            usecols = [usecols]

        # Filter the columns to keep
        cols_found = [col for col in df.columns for u in usecols if re.search(re.compile(u), col)]
        cols = ['Time'] + cols_found

        # If only the 'Time' column remains -> the columns specified do not exist
        if len(cols) == 1:
            raise ValueError(f"{bred}'{','.join(usecols)}'{reset}: column(s) not found.")
        
        # If at least one column of data is found
        else:
            # If not all the specified columns are found
            if len(cols) < len(usecols) + 1:
                cols_not_found = [col for col in df.columns if col not in cols]
                warnings.warn(f"{bred}'{','.join(cols_not_found)}'{reset}: column(s) not found.", UserWarning)

            # Filter the specified columns that are found
            df = df.loc[:, cols]
            print(f'Columns {bmag}{f"{reset}, {bmag}".join(df.columns[1:])}{reset} selected.')

    # If there are iterations to skip at the beginning or the end of the simulation
    if 'skipstart' in kwargs:
        skipstart = kwargs.get('skipstart')
        df = df.iloc[skipstart:,:]
        print(f"First {skipstart} iterations skipped.") # Verbose
    if 'skipend' in kwargs:
        skipend = kwargs.get('skipend')
        df = df.iloc[:-skipend, :]
        print(f"Last {skipend} iterations skipped.") # Verbose

    return df
        
# * ===================================================================================================
        
def _convert_numerical_data(df: pd.DataFrame) -> pd.DataFrame:

    # Progress bar setup
    tqdm.pandas(bar_format=cst.BAR_FORMAT,
                colour='green',
                ascii=' |',
                desc='Converting data to float')

    # If the 'Time' column contains strings representing integers
    if df.iloc[2, 0].isdigit():

        # The 'Time' column is converted to integer
        df['Time'] = df['Time'].apply(int)

        # Convert all the data to floats, except the 'Time' column
        df.iloc[:, 1:] = df.iloc[:, 1:].progress_apply(lambda x: x.astype(float))

    # If the 'Time' column contains strings representing floats, convert all the data to floats
    else:
        df = df.progress_apply(lambda x: x.astype(float))

    # Verbose
    print('Data converted.')

    return df

# * ===================================================================================================

def _files_to_df(file_paths: list, **kwargs) -> pd.DataFrame:

    out = _convert_numerical_data(
              _filter_data(
                  _remove_parenthesis(
                      _remove_NaN_cells(
                          _remove_NaN_columns(
                              _concat_data_files(file_paths)))), **kwargs))
    
    return out

# * ===================================================================================================

def _print_header(run_dirs: list) -> None:
    
    project = f'{bmag}{os.path.basename(cst.DEFAULT_DIR)}{bcyan}'
    runs_num = [os.path.basename(k) for k in run_dirs] 
    format_runs = f'{bmag}{f"{reset}, {bmag}".join(sorted(runs_num))}{bcyan}'
    title_df = pd.DataFrame({f'{reset}{bold}PROJECT{bcyan}': project,
                             f'{reset}{bold}RUN(S){bcyan}': format_runs},
                            index=['Data'])
    # Create a prettytable object
    pt = PrettyTable()

    for col in title_df.columns:
        pt.add_column(col, title_df[col].values)
        pt.align[col] = 'c'
        pt.min_width[col] = int(shutil.get_terminal_size().columns / 2) - 4

    # print the table
    print(bcyan)
    print(pt, end=f'{reset}')
    print('')
    
# %% ===================================================================================================

def _get_avg(df: pd.DataFrame, *,
             rng: int,
             type_avg: str = 'final',
             **kwargs) -> dict:

    # Select all the columns except 'Time' by default
    columns = list(df.columns)[1:]

    # If one or more columns are specified with 'usecols'
    if 'usecols' in kwargs:
        usecols = kwargs.get('usecols')
        columns = [col for col in list(df.columns)[1:] 
                   if re.search(re.compile(usecols), col)]
        
    # Return a dict of the mean value of each column over rng iterations
    if type_avg == 'final':
        return {c: df.loc[:, c].tail(rng).mean() for c in columns}
    
    # Get the window of series of observations of rng size for each column
    elif type_avg == 'moving':
        windows = {c: df.loc[:, c].rolling(rng) for c in columns}
        
        # Create a series of moving averages of each window for each column
        moving_avgs = {k: windows.get(k).mean().tolist() for k in windows}
        
        # Remove null entries
        final_dict = {k: moving_avgs.get(k)[rng - 1:] for k in moving_avgs}
        return final_dict

# %% ===================================================================================================

def _display_settings(**kwargs) -> tuple:
    # Set the default palette to 'hopium'
    matplotlib.rcParams['axes.prop_cycle'] = matplotlib.cycler(
        color=list(cst.HOPIUM_PALETTE.values()))
    
    # Set the color palette
    if 'palette' in kwargs:
        # 'classic' palette
        if kwargs.get('palette') == 'classic':
            matplotlib.rcParams['axes.prop_cycle'] = matplotlib.cycler(
                color='bgrcmyk')
        # 'cyberpunk' palette
        elif kwargs.get('palette') == 'cyberpunk':
            matplotlib.rcParams['axes.prop_cycle'] = matplotlib.cycler(
                color=list(cst.CYBERPUNK_PALETTE.values()))

    # Get the marker, space and underscore depending on LateX formatting
    marker, space, underscore = '', ' ', '_'

    if ('fancyplot' in kwargs or 'fp' in kwargs) \
    and (kwargs.get('fancyplot') or kwargs.get('fp')):
        #  LateX font and figure Args
        matplotlib.rcParams['text.usetex'] = True
        marker, space, underscore = '$', '\ ', '\_'
    else:
        matplotlib.rcParams['text.usetex'] = False
    sep = f'{space}{space}|{space}{space}'
    
    return marker, space, underscore, sep

# %% ===================================================================================================

def _get_data_from_run(run_path, *,
                       specdir: str = None,
                       graph_type: str = 'data',
                       probe: str = None,
                       **kwargs) -> list:
    
    run_id = os.path.basename(run_path)
    file_extension = '.dat'
    pp_dirs = []

    # Generic data
    if graph_type == 'data' and specdir != None:
        pp_dirs = _find_dirs(specdir, root_dir=run_path)
        error_dir = specdir

    # Residuals
    elif graph_type == 'residuals':
        pp_dirs = [os.path.join(run_path, 'postProcessing/residuals')]
        error_dir = "residuals"

    # Probes
    elif graph_type == 'probes' and probe != None:
        pp_dirs = [os.path.join(run_path, 'postProcessing/probes')]
        file_extension = probe
        error_dir = "probes"

    # Loop over the postpro dirs
    for pp in pp_dirs:

        # ! If wrong path
        if not os.path.isdir(pp):
            warnings.warn(f'No {bred}{error_dir}{reset} directory found.', UserWarning)

        # Get the list of file in a given run and the unique basename(s)
        file_paths = sorted(_find_files(file_extension, root_dir=pp))
        basenames = {os.path.basename(f) for f in file_paths}

        # ! More than one basename found
        if len(basenames) > 1:
            raise ValueError(f"More than one data type selected: {', '.join(bn for bn in basenames)}")
        
        # Yield a DataFrame and the run and postpro dir info
        df = _files_to_df(file_paths, **kwargs)
        if not df.empty:
            if file_extension == '.dat':
                yield {'run_id': run_id, 'pp_dir': os.path.basename(pp) , 'df': df}
            else:
                yield {'run_id': run_id, 'pp_dir': os.path.basename(file_paths[0]) , 'df': df}
        else:
            continue

# %% ===================================================================================================

def _get_unit(df,
              pp_dir,
              csv_df):
    
    # List of all the units for the pp dir in the csv
    unit_list = _get_postpro_labels(csv_df, pp_dir, "unit")
    unit_length = len(unit_list)

    # List of all the headers for the pp dir in the csv 
    header_list = _get_postpro_labels(csv_df, pp_dir, "postpro")

    # If at least one unit label
    if unit_length > 1:
        # The chosen unit for the graph is the first one in the list (skip the Time column)
        unit = [unit_list[i] for i in range(unit_length) if header_list[i] in df.columns][1]
    else:
        unit = None

    return unit

# %% ===================================================================================================

def _format_excel(file_path):
    # Load the Excel file
    workbook = load_workbook(file_path)
    sheet = workbook.active
    
    # Set font styles
    header_font = Font(name="Calibri", bold=True, size=14)
    content_font = Font(name="Calibri", size=12)
    
    # Set alignment
    alignment = Alignment(horizontal="center", vertical="center")
    
    # Set fill color
    fill_main = PatternFill(start_color="C7D1E0", end_color="C7D1E0", fill_type="solid")
    fill_data = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    
    # Set border
    border_color = "FF0000"
    thin_border = Border(top=Side(style=None), 
                         right=Side(style=None), 
                         bottom=Side(style=None), 
                         left=Side(style=None))
    
    # Format header row
    for cell in sheet[1]:
        cell.font = header_font
        cell.alignment = alignment
        cell.fill = fill_main
        cell.border = thin_border
        cell.value = cell.value.upper()
    
    # Format content rows
    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.font = content_font
            cell.alignment = alignment
            cell.fill = fill_data
            cell.border = thin_border
            
            if isinstance(cell.value, (int, float)):
                if cell.coordinate >= 'K':
                    cell.number_format = '0.00E+00'  # Scientific notation format code
    
    # Increase header row height
    sheet.row_dimensions[1].height = 40
    
    for row in sheet.iter_rows(min_row=2):
        sheet.row_dimensions[row[0].row].height = 20
    
    # Calculate the maximum text length in each column
    max_text_lengths = {}
    print(sheet.column_dimensions['G'].width)
    for row in sheet.iter_rows(min_row=1, values_only=True):
        for column_index, cell_value in enumerate(row, start=1):
            column_letter = get_column_letter(column_index)
            text_length = len(str(cell_value))
            if column_letter not in max_text_lengths or text_length > max_text_lengths[column_letter]:
                max_text_lengths[column_letter] = text_length

    # Set the column width as 1.2 times the maximum text length
    for column_letter, max_length in max_text_lengths.items():
        column_width = (max_length * 1.2) + 2  # Add some extra padding
        sheet.column_dimensions[column_letter].width = column_width
    
    # Save the modified Excel file
    workbook.save(file_path)