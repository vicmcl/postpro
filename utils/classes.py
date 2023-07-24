import importlib
import matplotlib
import matplotlib.pyplot as plt
import numpy as np
import os
import pandas as pd
import re
import seaborn as sns
import shutil
import utils.find as fd

from datetime import timedelta
from file_read_backwards import FileReadBackwards
from functools import partial
from getpass import getuser
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from pathlib import Path
from prettytable import PrettyTable
from socket import gethostname
from tqdm import tqdm
from warnings import warn

from openpyxl.styles import (
    Font,
    PatternFill,
    Alignment,
    Border,
    Side,
)

from scipy.fft import (
    fft,
    fftfreq,
)

from statistics import (
    stdev,
    mean,
)
 
class Run():
    
    def __init__(self, run_path) -> None:
        # ! If no run path
        if not run_path.exists():
            raise ValueError("Run path does not exist.")
        
        pp_path: Path = run_path / "postProcessing"
        system_path: Path = run_path / "system"
        constant_path: Path = run_path / "constant" 
        
        # ! If no postProcessing dir
        if not all(p.exists() for p in [pp_path, system_path, constant_path]):
            raise ValueError("Missing either postProcessing, system of constant dir inside this run.")
        
        self.path: Path = run_path
        self.postpro_dirs: list[Path] = [d for d in pp_path.iterdir() if d.is_dir()]
        self.project: Path = run_path.parent
        
    info: str
    
class PostProDir():
    
    def __init__(self, *, name: str, run: Run) -> None:
        pp_path: Path = run.path / "postProcessing" / name
        
        # ! If no dir
        if not pp_path.exists():
            raise ValueError(f'No dir named "{name}" inside postProcessing.')
        
        self.path: Path = pp_path
        self.timesteps: list[str] = [t.name for t in self.path.iterdir() if t.is_dir()]
        self.project: Path = run.project
        self.run: Path = run.path
    
class PostProFile():
    
    def __init__(self, *, name: str, post: PostProDir, ts: str) -> None:
        
        # Find the file path based on the pp dir, the timestep and the name regex
        fpath: Path = fd.find_files(name, root_dir=post.path / ts)
        
        # ! If multiple files
        if len(fpath) > 1:
            raise ValueError("Ambiguous name, mutiple files found.")
        # ! If no file
        if not fpath[0].exists():
            raise ValueError("File not found.")
        
        self.path: Path = fpath[0]
        self.labels: dict = fd.label_names(self.path)
        self.run: Path = post.run
        self.project: Path = post.project
        self.post: Path = post.path
        self.timestep: str = ts
        
    def __str__(self):
        return f"Name:\t\t{self.path.name}\n \
               \rProject:\t{self.project.name}\n \
               \rRun:\t\t{self.run.name}\n \
               \rPost Dir:\t{self.post.name}\n \
               \rTimestep:\t{self.timestep}\n \
               \rLabels:\t\t{', '.join(lab for lab in self.labels.get('file_labels'))}"

    
    
    

    
    
    
    
    
    