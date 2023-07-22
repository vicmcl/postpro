import importlib
import matplotlib
import matplotlib.pyplot as plt
import numpy as np
import os
import pandas as pd
import re
import seaborn as sns
import shutil

from datetime import timedelta
from file_read_backwards import FileReadBackwards

from functools import partial
from getpass import getuser
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from pathlib import Path
from prettytable import PrettyTable
from scipy.fft import fft, fftfreq
from socket import gethostname
from statistics import stdev, mean
from tqdm import tqdm
from warnings import warn

import utils.find as find
import utils.fetch as fetch
import utils.constants as cst

# * ===================================================================================================

def issteady(run: str) -> bool:
    log_file = find.find_logs(find.find_runs(run)[0])[0]
    with FileReadBackwards(log_file) as frb:
        for line in frb:
            if line.startswith("Time ="):
                if line.split()[-1].isdigit():
                    return True
                else:
                    return False
                
# * ===================================================================================================

def ncol(handles: list) -> int:

    max_text_length = 60
    nhandles = len(handles)
    total_length = sum(len(text) for text in handles) + 3 * nhandles
    
    if total_length > max_text_length:
        if nhandles > 6:
            ncol = 4
        else:
            ncol = max(int(nhandles / 2), int((nhandles + 1) / 2))
        
        row_index = range(0, nhandles - ncol, ncol)
        for i in row_index:
            words_in_row = [handles[k] for k in range(i, i + ncol)]
            if sum(len(word) for word in words_in_row) > max_text_length:
                ncol -= 1
                break
    else:
        ncol = nhandles
    return ncol

# * ===================================================================================================

def _print_header(run_dirs: list[Path]) -> None:
    
    # Check project is unique
    if not len({r.parent.name for r in run_dirs}) == 1:
        raise ValueError("Multiple project directories found.")
    
    # If unique project
    project = f'{cst.bmag}{run_dirs[0].parent.name}{cst.cst.bcyan}'
    runs_num = [k.name for k in run_dirs] 
    format_runs = f'{cst.bmag}{f"{cst.reset}, {cst.bmag}".join(sorted(runs_num))}{cst.bcyan}'
    title_df = pd.DataFrame({f'{cst.reset}{cst.bold}PROJECT{cst.bcyan}': project,
                             f'{cst.reset}{cst.bold}RUN(S){cst.bcyan}': format_runs},
                            index=['Data'])
    # Create a prettytable object
    pt = PrettyTable()

    for col in title_df.columns:
        pt.add_column(col, title_df[col].values)
        pt.align[col] = 'c'
        pt.min_width[col] = int(shutil.get_terminal_size().columns / 2) - 4

    # print the table
    print(cst.bcyan)
    print(pt, end=f'{cst.reset}')
    print('')
    
# * ===================================================================================================

def get_avg(df: pd.DataFrame, *,
             rng: int,
             type_avg: str = 'final',
             **kwargs) -> dict:

    # Select all the columns except 'Time' by default
    columns = list(df.columns)[1:]

    # If one or more columns are specified with 'usecols'
    if 'usecols' in kwargs:
        usecols = kwargs.get('usecols')
        columns = [col for col in list(df.columns)[1:] 
                   if re.search(re.compile(usecols), col)]
        
    # Return a dict of the mean value of each column over rng iterations
    if type_avg == 'final':
        return {c: df.loc[:, c].tail(rng).mean() for c in columns}
    
    # Get the window of series of observations of rng size for each column
    elif type_avg == 'moving':
        windows = {c: df.loc[:, c].rolling(rng) for c in columns}
        
        # Create a series of moving averages of each window for each column
        moving_avgs = {k: windows.get(k).mean().tolist() for k in windows}
        
        # Remove null entries
        final_dict = {k: moving_avgs.get(k)[rng - 1:] for k in moving_avgs}
        return final_dict
    
# * ===================================================================================================

def _format_excel(file_path):
    # Load the Excel file
    workbook = load_workbook(file_path)
    sheet = workbook.active
    
    # Set font styles
    header_font = Font(name="Calibri", bold=True, size=14)
    content_font = Font(name="Calibri", size=12)
    
    # Set alignment
    alignment = Alignment(horizontal="center", vertical="center")
    
    # Set fill color
    fill_main = PatternFill(start_color="C7D1E0", end_color="C7D1E0", fill_type="solid")
    fill_data = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    
    # Set border
    border_color = "FF0000"
    thin_border = Border(top=Side(style=None), 
                         right=Side(style=None), 
                         bottom=Side(style=None), 
                         left=Side(style=None))
    
    # Format header row
    for cell in sheet[1]:
        cell.font = header_font
        cell.alignment = alignment
        cell.fill = fill_main
        cell.border = thin_border
        cell.value = cell.value.upper()
    
    # Format content rows
    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.font = content_font
            cell.alignment = alignment
            cell.fill = fill_data
            cell.border = thin_border
            
            if isinstance(cell.value, (int, float)):
                if cell.coordinate >= 'K':
                    cell.number_format = '0.00E+00'  # Scientific notation format code
    
    # Increase header row height
    sheet.row_dimensions[1].height = 40
    
    for row in sheet.iter_rows(min_row=2):
        sheet.row_dimensions[row[0].row].height = 20
    
    # Calculate the maximum text length in each column
    max_text_lengths = {}
    print(sheet.column_dimensions['G'].width)
    for row in sheet.iter_rows(min_row=1, values_only=True):
        for column_index, cell_value in enumerate(row, start=1):
            column_letter = get_column_letter(column_index)
            text_length = len(str(cell_value))
            if column_letter not in max_text_lengths or text_length > max_text_lengths[column_letter]:
                max_text_lengths[column_letter] = text_length

    # Set the column width as 1.2 times the maximum text length
    for column_letter, max_length in max_text_lengths.items():
        column_width = (max_length * 1.2) + 2  # Add some extra padding
        sheet.column_dimensions[column_letter].width = column_width
    
    # Save the modified Excel file
    workbook.save(file_path)