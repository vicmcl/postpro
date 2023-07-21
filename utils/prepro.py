import constants as cst
import importlib
import matplotlib
import matplotlib.pyplot as plt
import numpy as np
import os
import pandas as pd
import re
import seaborn as sns
import shutil

from datetime import timedelta
from file_read_backwards import FileReadBackwards

from functools import partial
from getpass import getuser
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from pathlib import Path
from prettytable import PrettyTable
from scipy.fft import fft, fftfreq
from socket import gethostname
from statistics import stdev, mean
from tqdm import tqdm
from warnings import warn

import utils.find as find
import utils.fetch as fetch

# * ===================================================================================================

def _concat_data_files(file_paths: list[Path]) -> tuple:
        
        data_list = []

        # Sort file paths 
        timesteps = [np.float64(f.parent.name) for f in file_paths]
        sorted_file_paths = [x for _, x in sorted(zip(timesteps, file_paths), key=lambda pair: pair[0])]

        # Get columns name
        cols = find.label_names(sorted_file_paths[0])['postpro_labels']

        # Loop over the file paths
        for fpath in sorted_file_paths:
            pp_dir = fpath.parents[1].name
            timestep = fpath.parent.name

            # Parse file
            with fpath.open() as f:
                for line in f:
                    if line.startswith('# Time'):
                        break

                # Verbose
                if 'probes' in fpath.parents:
                    print(f'Parsing {bmag}probe {fpath.name}{reset} at timestep {bmag}{timestep}{reset}:')
                else:
                    print(f'Parsing {bmag}{pp_dir}{reset} at timestep {bmag}{timestep}{reset}:')
                
                # Set progress bar
                pbar = tqdm(total=0, unit=' lines')

                # Split each line to store the data in a tuple
                for line in f:
                    data = tuple(line.split())
                    data_list.append(data)
                    pbar.update(1)
                pbar.close()

                #Verbose
                print('File parsed.')
        
        # If at least 2 files are concatenated, display their start timestep
        if len(file_paths) > 1:
            fmt_sep = f"{reset}, {bmag}"
            fmt_timesteps = f"{fmt_sep}".join(sorted([str(i) for i in timesteps]))
            print(f'Concatenated files at timesteps {bmag}{fmt_timesteps}{reset}.')

        return data_list, cols

# * ===================================================================================================

def _remove_NaN_columns(concat_data: tuple) -> pd.DataFrame:

    data, headers = concat_data
        
    # Check for NaN columns to remove
    # Initialize the indexes of headers and columns to drop
    idx_headers_to_drop = []
    idx_cols_to_drop = []

    # DataFrame created without the headers
    df = pd.DataFrame(data)
    
    # Loop over the columns of the DataFrame
    for i in tqdm(range(len(df.columns)),
                    bar_format=cst.BAR_FORMAT,
                    ascii=' |',
                    colour='green',
                    desc='Checking for NaN columns'):
        
        # If a column filled with NaN is found    
        if df.iloc[:, i][df.iloc[:, i] != 'N/A'].size == 0:

            # Its index is saved to remove the corresponding header and column
            idx_cols_to_drop.append(i)
            idx_headers_to_drop.append(i)

            # If the velocity vector is NaN, 
            # the Ux, Uy and Uz headers must be removed
            if 'Uy' in headers[i+1] or 'Uz' in headers[i+1]:
                idx_headers_to_drop.append(i+1)
                if 'Uz' in headers[i+2]:
                    idx_headers_to_drop.append(i+2)

    # Verbose
    if idx_cols_to_drop:
        print('NaN column(s) removed.')
    else:
        print('No column to remove.')

    # Associate the columns with their headers
    df = df.drop(idx_cols_to_drop, axis=1)
    headers = [val for i, val in enumerate(headers) if i not in idx_headers_to_drop]
    df.columns = headers
    
    return df 

# * ===================================================================================================
    
def _remove_NaN_cells(df: pd.DataFrame) -> pd.DataFrame:

    # Check for NaN cells to remove
    tqdm.pandas(bar_format=cst.BAR_FORMAT,
                ascii=' |',
                colour='green',
                desc='Checking for remaining NaN cells')
    
    # Save the initial number of rows in the DataFrame
    previous_size = df.size

    # Drop the rows with one or more NaN or None cells
    df = df[~df.progress_apply(lambda x: x.isin(['N/A', None])).any(axis=1)]

    # Verbose
    if df.size != previous_size:
        print('NaN cell(s) removed.')
    else:
        print('No cell to remove.')

    return df

# * ===================================================================================================

def _remove_parenthesis(df: pd.DataFrame) -> pd.DataFrame:

    # Check for parenthesis to remove
    cols_with_parenthesis = [col for col in df.columns if df[col][1].startswith('(') or df[col][1].endswith(')')]

    # If one or more columns contain cells with parenthesis
    if len(cols_with_parenthesis) > 0:

        # Progress bar setup
        tqdm.pandas(bar_format=cst.BAR_FORMAT,
                    colour='green',
                    ascii=' |',
                    desc='Formatting vectors')
        
        # Remove the parenthesis in the identified columns
        df[cols_with_parenthesis] = df[cols_with_parenthesis].progress_apply(lambda x: x.str.replace(r'[()]', '', regex=True))

        # Verbose
        print('Vectors formatted.')

    return df

# * ===================================================================================================

def _filter_data(df: pd.DataFrame, **kwargs) -> pd.DataFrame:

    # If one or more columns are specified by the user
    if 'usecols' in kwargs:
        usecols = kwargs.get('usecols')
        if isinstance(usecols, str):
            usecols = [usecols]

        # Filter the columns to keep
        cols_found = [col for col in df.columns for u in usecols if re.search(re.compile(u), col)]
        cols = ['Time'] + cols_found

        # If only the 'Time' column remains -> the columns specified do not exist
        if len(cols) == 1:
            raise ValueError(f"{bred}'{','.join(usecols)}'{reset}: column(s) not found.")
        
        # If at least one column of data is found
        else:
            # If not all the specified columns are found
            if len(cols) < len(usecols) + 1:
                cols_not_found = [col for col in df.columns if col not in cols]
                warn(f"{bred}'{','.join(cols_not_found)}'{reset}: column(s) not found.", UserWarning)

            # Filter the specified columns that are found
            df = df.loc[:, cols]
            print(f'Columns {bmag}{f"{reset}, {bmag}".join(df.columns[1:])}{reset} selected.')

    # If there are iterations to skip at the beginning or the end of the simulation
    if 'skipstart' in kwargs:
        skipstart = kwargs.get('skipstart')
        df = df.iloc[skipstart:,:]
        print(f"First {skipstart} iterations skipped.") # Verbose
    if 'skipend' in kwargs:
        skipend = kwargs.get('skipend')
        df = df.iloc[:-skipend, :]
        print(f"Last {skipend} iterations skipped.") # Verbose

    return df
        
# * ===================================================================================================
        
def _convert_numerical_data(df: pd.DataFrame) -> pd.DataFrame:

    # Progress bar setup
    tqdm.pandas(bar_format=cst.BAR_FORMAT,
                colour='green',
                ascii=' |',
                desc='Converting data to float')

    # If the 'Time' column contains strings representing integers
    if df.iloc[2, 0].isdigit():

        # The 'Time' column is converted to integer
        df['Time'] = df['Time'].apply(int)

        # Convert all the data to floats, except the 'Time' column
        df.iloc[:, 1:] = df.iloc[:, 1:].progress_apply(lambda x: x.astype(np.float64))

    # If the 'Time' column contains strings representing floats, convert all the data to floats
    else:
        df = df.progress_apply(lambda x: x.astype(np.float64))

    # Verbose
    print('Data converted.')

    return df

# * ===================================================================================================

def _files_to_df(file_paths: list, **kwargs) -> pd.DataFrame:

    out = _convert_numerical_data(
              _filter_data(
                  _remove_parenthesis(
                      _remove_NaN_cells(
                          _remove_NaN_columns(
                              _concat_data_files(file_paths)))), **kwargs))
    
    return out