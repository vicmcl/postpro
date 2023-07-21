import constants as cst
import importlib
import matplotlib
import matplotlib.pyplot as plt
import numpy as np
import os
import pandas as pd
import re
import seaborn as sns
import shutil

from datetime import timedelta
from file_read_backwards import FileReadBackwards

from functools import partial
from getpass import getuser
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from pathlib import Path
from prettytable import PrettyTable
from scipy.fft import fft, fftfreq
from socket import gethostname
from statistics import stdev, mean
from tqdm import tqdm
from warnings import warn

import utils.find as find

# * ===================================================================================================

def fetch_run_data(run_path: Path, *,
                       specdir: str = None,
                       probe: str = None,
                       **kwargs) -> list:
    
    run_id = run_path.name
    file_extension = '.dat'
    pp_dirs = []

    if probe != None and specdir != None:
        raise ValueError('probe and specdir are mutually exclusive.')

    # Generic data
    if specdir != None:
        pp_dirs = find.find_dirs(specdir, root_dir=run_path)
        error_dir = specdir

    # Probes
    elif probe != None:
        pp_dirs = [run_path / 'postProcessing/probes']
        file_extension = probe
        error_dir = "probes"
    
    # Residuals
    else:
        pp_dirs = [run_path / 'postProcessing/residuals']
        error_dir = "residuals"

    # Loop over the postpro dirs
    for pp in pp_dirs:

        # ! If wrong path
        if not pp.is_dir():
            warn(f'No {bred}{error_dir}{reset} directory found.', UserWarning)

        # Get the list of file in a given run and the unique basename(s)
        file_paths = sorted(find.find_files(file_extension, root_dir=pp))
        basenames = {f.name for f in file_paths}

        # ! More than one basename found
        if len(basenames) > 1:
            raise ValueError(f"More than one data type selected: {', '.join(bn for bn in basenames)}")
        
        # Yield a DataFrame and the run and postpro dir info
        df = _files_to_df(file_paths, **kwargs)
        if not df.empty:
            if file_extension == '.dat':
                yield {'run_id': run_id, 'pp_dir': pp.name , 'df': df}
            else:
                yield {'run_id': run_id, 'pp_dir': file_paths[0].name , 'df': df}
        else:
            continue

# * ===================================================================================================

def fetch_unit(*, df,
              csv_df,
              pp_dir,
              probe,
              **kwargs):
    
    if probe != None:
        if 'unit' in kwargs: unit = kwargs.get("unit")
        elif probe == "p" or probe =="^p" or probe == "p$": unit = 'Pa'
        elif probe == "k" or probe =="^k" or probe == "k$": unit = 'J/kg'
        else: unit = None
    
    elif pp_dir == 'residuals':
        unit = 'Residuals'
        
    else:
        # List of all the units for the pp dir in the csv
        unit_list = find.get_labels(csv_df, pp_dir, "unit")
        unit_length = len(unit_list)

        # List of all the headers for the pp dir in the csv 
        header_list = find.get_labels(csv_df, pp_dir, "postpro")

        # If at least one unit label
        if unit_length > 1:
            # The chosen unit for the graph is the first one in the list (skip the Time column)
            unit = [unit_list[i] for i in range(unit_length) if header_list[i] in df.columns][1]
        else:
            unit = None

    return unit